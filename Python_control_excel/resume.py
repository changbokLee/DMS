from openpyxl import Workbook # 기본워크시트만들때
from openpyxl.workbook import Workbook 
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import (PieChart, Reference) 
from openpyxl.chart.series import DataPoint

# resume 헤드라인 작성
wb = Workbook()
ws = wb.active
ws.title ="Resume"
ws.merge_cells('A1:H1')
ws['A1'] = 'Resume'
ca1 = ws['A1']
ca1.font = Font(name ='맑은고딕', size='30', bold='True', color='FF000000')
box = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), 
top=Side(border_style='thin', color='FF000000'),bottom=Side(border_style='thin', color='FF000000'), diagonal=Side(border_style='thin', color='FF000000'),
diagonal_direction=0, outline=Side(border_style='thin', color='FF000000'), vertical=Side(border_style='thin', color='FF000000'), horizontal=Side(border_style='thin', 
color='FF000000'))
ca1.border = box
ca1.fill = PatternFill(patternType='solid', fgColor= 'FFC000')
ca1.alignment = Alignment(horizontal='center',vertical='center')

# 테이블 이력서  내용 , 리스트사용 
about_me = [
    ['Name', 'LeeChangbok', 'School', 'Chung-nam-college','Phone', '010 -1234-1234'],
    ['Age', 25, 'Major', 'Developer','Sns','Facebook'],
    ['Gender', 'Male', 'address', 'snackfor','company','snakcfor'],
    
]
ws.append(['Contents','Atribute',"Content1","Atributes","Content2","Atribute3"])
for row in about_me:
    ws.append(row)

tab = Table(displayName ="WhoAmI", ref = "A2:E4")
style = TableStyleInfo(name = "resume", showFirstColumn= False, showLastColumn= False,
showColumnStripes= True)
tab.tableStyleInfo = style
ws.add_table(tab)


# 스킬차트 만들기 ,리스트 사용
skills = [ 
        ['language', 'Skills'],
        ['Python', 50],
        ['Html', 20],
        ['Css', 10],             
        ['Django', 30],
]

for rows in skills:
    ws.append(rows)


chart = PieChart()
chart.title = "Skill-Chart"
labels = Reference(ws, min_col=1, min_row=2, range_string='resume!A6:A10' )
data = Reference(ws, min_col=1, min_row=2, max_row=5, range_string= 'resume!B6:B10')
chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)

ws.add_chart(chart, "A15")

# 저장
wb.save('resume.xlsx')
